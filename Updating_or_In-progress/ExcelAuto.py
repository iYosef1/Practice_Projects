import openpyxl
from openpyxl.worksheet.table import Table  #print(ws.tables.items())
from openpyxl import Workbook, load_workbook
import pprint

#wb = load_workbook('ExcelAutoTestWorkbook.xlsx')


def all_sheets(workbook):
    return(workbook.sheetnames)

def access_sheet(workbook, sheet_name):
    worksheet = workbook[sheet_name]

def new_sheet(workbook, sheet_name):
    workbook.create_sheet(sheet_name)

    
def std_data_extract(worksheet, heading, coordinates):
    """
    This function extracts every cell's input from a standard
    (uninserted) table of x number of columns and y number of
    rows from an (existing) workbook. The data can be extracted
    by column or row headings by passing in "column" or "row"
    into the heading parameter. The argument for coordinates
    requires a tuple referencing the top left cell first and the
    the bottom right cell second as the range of the table.
    """
    data = dict()
    table_dimension = worksheet[coordinates[0]:coordinates[1]]
    if heading == "column":
        column_headers = table_dimension[0]
        for header, item_num in zip(column_headers, range(len(column_headers))):
            column_content = []
            for row in table_dimension[1:]:
                column_content.append(row[item_num].value)
            data[header.value] = column_content
        #pprint.pprint(data)     

    elif heading == "row":
        row_content = []
        for row in table_dimension:
            row_content = []
            for item in row[1:]:     
                row_content.append(item.value)
            data[row[0].value] = row_content
        #pprint.pprint(data)

    return data




'''
ws = wb.active
row = ws['A1':'C4']

col = std_data_extract(ws, "row", ('A1','C4'))



print('\n\n')
for x in row:
    print(x[0].value, x[1].value, x[2].value)
    pass
'''



#wb.save('ExcelAutoTestWorkbook.xlsx')
